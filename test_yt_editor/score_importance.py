"""Read a table of transcript segments (index/start/end/content) from a CSV
or Excel file, ask Gemini to score each segment's importance (0.0-1.0), and
write the result back with a new score column.

Usage:
    python score_importance.py --input segments.xlsx
    python score_importance.py --input segments.csv --output scored.csv
    python score_importance.py --input segments.xlsx --batch-size 20

Input file columns (Korean or English headers both work):
    인덱스/index, 시작/start, 끝/end, 내용/content

Setup:
    pip install -r requirements.txt
    .env 파일에 GEMINI_API_KEY="<API_KEY>" 작성 (https://aistudio.google.com/api-keys)
"""
import argparse
import json
import os
import re
import time
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()

GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")

_COLUMN_ALIASES = {
    "index": {"index", "인덱스", "idx"},
    "start": {"start", "시작"},
    "end": {"end", "끝"},
    "content": {"content", "내용", "text", "script"},
}

SYSTEM_PROMPT = """\
You are a professional video editor's assistant, scoring transcript segments
of a long-form video for how important each one is to a highlight reel.
Segments are short chunks of speech, not full scenes - always judge a segment
using its own content PLUS the surrounding segments, never in isolation.

What raises a score: viewer-reaction moments (funny, shocking, success/
failure), high information density (key explanations, conclusions, new
facts), narrative turning points (start/end of a challenge, conflict and
resolution), content that represents the video's core theme/title, emotional
peaks (candid confessions, moving moments).

What lowers a score: silence, filler ("um...", throat-clearing), repeated
content, dead time (camera setup, waiting, moving locations), tangents
unrelated to the main content.

Score bands:
- 0.00-0.15 silence/filler, safe to always cut
- 0.16-0.35 small talk or low-value content
- 0.36-0.55 needed for context but not highlight-worthy on its own
- 0.56-0.75 engaging: funny, informative, a minor twist
- 0.76-0.90 quite important, likely to be referenced by viewers/comments
- 0.91-1.00 must-keep: the moment this part of the video is about

Context rules: a short plain sentence can still score high if it is the
payoff/conclusion of a longer setup in prior segments - do not undervalue it
just because it is short. Segments belonging to the same episode/event
(setup -> escalation -> payoff) should carry similar, gradually-changing
scores; a score should only jump sharply when the topic itself changes
(e.g. into an unrelated tangent).

Return ONLY a JSON array, no markdown fences, one entry per segment in the
same order as given, using each segment's own "row" number:
[{"row": 0, "score": 0.8, "reason": "short reason"}, {"row": 1, "score": 0.1, "reason": "short reason"}]
"reason" is one short phrase for a human to audit later; it is not used by
downstream code, so keep it brief.
"""


# --- Gemini call (with retry on transient errors) --------------------------

_RETRYABLE_MARKERS = ("503", "UNAVAILABLE", "429", "RESOURCE_EXHAUSTED", "500", "overloaded")
_RETRY_DELAYS = [10, 30, 60]


def _chat(system: str, user: str) -> str:
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise ValueError("환경 변수 'GEMINI_API_KEY'가 설정되지 않았습니다. .env 파일을 확인해주세요.")

    from google import genai
    from google.genai import types

    client = genai.Client(api_key=api_key)

    last_exc = None
    for attempt, delay in enumerate([0] + _RETRY_DELAYS):
        if delay:
            print(f"[재시도 {attempt}/{len(_RETRY_DELAYS)}] {delay}초 후 재시도: {last_exc}")
            time.sleep(delay)
        try:
            response = client.models.generate_content(
                model=GEMINI_MODEL,
                contents=user,
                config=types.GenerateContentConfig(system_instruction=system, temperature=0.3),
            )
            return response.text or ""
        except Exception as exc:
            last_exc = exc
            if not any(m in str(exc) for m in _RETRYABLE_MARKERS):
                break
    raise RuntimeError(f"Gemini API 호출 실패: {last_exc}") from last_exc


def _extract_json_array(text: str) -> list:
    text = text.strip()
    fenced = re.search(r"```(?:json)?\s*(\[.*?\])\s*```", text, re.DOTALL)
    if fenced:
        text = fenced.group(1)
    else:
        bracket = re.search(r"\[.*\]", text, re.DOTALL)
        if bracket:
            text = bracket.group(0)
    data = json.loads(text)
    if not isinstance(data, list):
        raise ValueError("응답이 JSON 배열 형식이 아닙니다.")
    return data


# --- table I/O ---------------------------------------------------------------

def _resolve_columns(header: list) -> dict:
    """Map our canonical field names to whatever column names the file uses."""
    normalized = {h.strip().lower(): h for h in header}
    resolved = {}
    for field, aliases in _COLUMN_ALIASES.items():
        match = next((normalized[a] for a in aliases if a in normalized), None)
        if match is None:
            raise ValueError(f"'{field}' 컬럼을 찾을 수 없습니다. 헤더: {header}")
        resolved[field] = match
    return resolved


def read_rows(path: Path) -> list:
    """Load rows as a list of dicts with keys: index, start, end, content."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        import openpyxl

        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(c) if c is not None else "" for c in rows[0]]
        cols = _resolve_columns(header)
        out = []
        for r in rows[1:]:
            record = dict(zip(header, r))
            out.append({
                "index": record[cols["index"]],
                "start": record[cols["start"]],
                "end": record[cols["end"]],
                "content": record[cols["content"]],
            })
        return out

    import csv

    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        cols = _resolve_columns(reader.fieldnames or [])
        return [
            {
                "index": row[cols["index"]],
                "start": row[cols["start"]],
                "end": row[cols["end"]],
                "content": row[cols["content"]],
            }
            for row in reader
        ]


def write_rows(path: Path, rows: list) -> None:
    fieldnames = ["index", "start", "end", "content", "importance_score"]

    if path.suffix.lower() in (".xlsx", ".xls"):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(fieldnames)
        for row in rows:
            ws.append([row[f] for f in fieldnames])
        wb.save(path)
        return

    import csv

    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


# --- scoring -----------------------------------------------------------------

def _build_batch_prompt(rows: list) -> str:
    lines = [
        f"row {i}: [{r['start']} - {r['end']}] {r['content']}"
        for i, r in enumerate(rows)
    ]
    return "\n".join(lines)


def _parse_scores(raw_text: str, batch_size: int) -> dict:
    data = _extract_json_array(raw_text)
    scores = {}
    for item in data:
        if not isinstance(item, dict):
            continue
        try:
            row = int(item["row"])
            score = float(item["score"])
        except (KeyError, TypeError, ValueError):
            continue
        if 0 <= row < batch_size:
            scores[row] = max(0.0, min(1.0, score))
    return scores


def score_batch(rows: list) -> list:
    """Ask Gemini to score one batch of rows together (shared context)."""
    if not rows:
        return []
    content = _chat(SYSTEM_PROMPT, _build_batch_prompt(rows))
    scores = _parse_scores(content, len(rows))
    return [scores.get(i, 0.0) for i in range(len(rows))]


def score_all(rows: list, batch_size: int) -> list:
    for start in range(0, len(rows), batch_size):
        batch = rows[start:start + batch_size]
        scores = score_batch(batch)
        for row, score in zip(batch, scores):
            row["importance_score"] = score
        print(f"[채점 완료] {min(start + batch_size, len(rows))}/{len(rows)}")
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--input", required=True, help="path to a .csv or .xlsx file")
    parser.add_argument("--output", help="output path (default: <input>_scored.<ext>)")
    parser.add_argument("--batch-size", type=int, default=30,
                        help="segments per LLM call, for shared context (default: 30)")
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output) if args.output else (
        input_path.with_stem(input_path.stem + "_scored")
    )

    rows = read_rows(input_path)
    if not rows:
        print("입력 파일에 데이터가 없습니다.")
        return

    rows = score_all(rows, args.batch_size)
    write_rows(output_path, rows)
    print(f"\n✅ 완성! {len(rows)}개 구간 채점 -> {output_path}")


if __name__ == "__main__":
    main()
